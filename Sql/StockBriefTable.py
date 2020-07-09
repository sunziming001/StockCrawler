from Sql.Connect import get_sql_conn
import openpyxl
import pandas as pd


class StockBrief:
    code_id = ''
    name = ''


class StockBriefTable:

    @staticmethod
    def init_stock_brief_from_cvs(cvs_path):
        StockBriefTable.clear_brief_table()
        df = pd.read_csv(cvs_path)
        total_rows = df.shape[0]
        brief_list = []
        for i in range(0, total_rows):
            brief = StockBrief()
            str_progress = str(i) + "/" + str(total_rows)
            print("\b" * (len(str_progress) * 2), end="")
            print(str_progress, end="")

            brief.code_id = df.iloc[i][1]
            brief.name = df.iloc[i][1]
            StockBriefTable.insert_stock_brief(brief)



    @staticmethod
    def init_stock_brief_from_xl(xl_path):
        StockBriefTable.clear_brief_table()
        wb = openpyxl.load_workbook(xl_path)
        sheet_names = wb.sheetnames
        default_sheet = wb[sheet_names[0]]
        max_row = default_sheet.max_row

        for i in range(2, max_row):
            brief = StockBrief()
            str_progress = str(i) + "/" + str(max_row)
            print("\b" * (len(str_progress) * 2), end="")
            print(str_progress, end="")
            brief.code_id = default_sheet.cell(i, 1).value
            brief.name = default_sheet.cell(i, 2).value
            StockBriefTable.insert_stock_brief(brief)

    @staticmethod
    def clear_brief_table():
        conn = get_sql_conn()
        sql = 'delete from StockBrief'
        conn.execute(sql)
        conn.commit()

    @staticmethod
    def insert_stock_brief(stock_brief):
        conn = get_sql_conn()
        sql = ('insert into StockBrief (codeId, name) values ('
               + '\'' + stock_brief.code_id + '\'' + ", "
               + '\'' + stock_brief.name + '\'' + ");")
        conn.execute(sql)
        conn.commit()


    @staticmethod
    def get_stock_id_list():
        ret_list = []
        conn = get_sql_conn()
        sql = 'select codeId from StockBrief order by codeId;'
        cursor = conn.execute(sql)
        for row in cursor:
            code_id = row[0]
            ret_list.append(code_id)
        return ret_list
